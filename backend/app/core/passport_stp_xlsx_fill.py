"""Заполнение шаблона паспорта ВЛ (СТП) из снимка ЛЭП — как в образце stp_passport_vl.xlsx."""

from __future__ import annotations

import json
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "stp_passport_vl.xlsx"


def _ru_decimal(x: Any) -> str:
    if x is None:
        return ""
    try:
        s = f"{float(x):.5f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return str(x)
    return s.replace(".", ",")


def _find_row_contains(ws, text: str, max_row: int = 500) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=1):
        c = row[0]
        v = c.value
        if v is not None and text in str(v):
            return c.row
    return None


def _find_table_25_bounds(ws) -> Tuple[int, int]:
    """Строки данных таблицы 2.5: с .. по (не включая строку ИТОГО)."""
    t = _find_row_contains(ws, "Таблица 2.5")
    if not t:
        return 225, 244
    # Заголовок, строка «№ п/п» / «Тип опоры», строка номеров колонок 1,2,3 — затем данные
    data_start = t + 3
    total_r = None
    for row in ws.iter_rows(min_row=data_start, max_row=data_start + 300, min_col=1, max_col=1):
        c = row[0]
        v = c.value
        if v is not None and "ИТОГО" in str(v).upper() and "ОПОР" in str(v).upper():
            total_r = c.row
            break
    if total_r is None:
        return data_start, data_start + 50
    return data_start, total_r - 1


def _clear_table_25_block(ws, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for col in (1, 4, 12):
            ws.cell(r, col).value = None


def build_stp_line_passport_xlsx(
    snapshot_data: Dict[str, Any],
    title: str,
    stp_reference: Optional[str],
    manual_sections: Optional[Dict[str, Any]],
) -> bytes:
    """Собирает XLSX по шаблону СТП для ЛЭП; при отсутствии шаблона — RuntimeError."""
    if not TEMPLATE_PATH.is_file():
        raise RuntimeError(f"Шаблон паспорта не найден: {TEMPLATE_PATH}")

    bio = BytesIO(TEMPLATE_PATH.read_bytes())
    wb = load_workbook(bio)
    ws = wb["TDSheet"]

    pl = snapshot_data.get("power_line") or {}
    name = str(pl.get("name") or "").strip()
    v = pl.get("voltage_level_kv")
    try:
        vk = int(float(v)) if v is not None else None
    except (TypeError, ValueError):
        vk = None

    if vk is not None and name:
        ws["A10"] = f"ВОЗДУШНОЙ ЛИНИИ ВЛ-{vk} кВ {name}"
    elif name:
        ws["A10"] = f"ВОЗДУШНОЙ ЛИНИИ {name}"
    else:
        ws["A10"] = title or "ВОЗДУШНОЙ ЛИНИИ"

    ss1 = pl.get("substation_start") or {}
    ss2 = pl.get("substation_end") or {}
    parts: List[str] = []
    if ss1.get("name"):
        parts.append(f"ПС {ss1['name']}")
    if ss2.get("name"):
        parts.append(f"— ПС {ss2['name']}")
    if parts:
        ws["A11"] = "от подстанции  " + " ".join(parts)

    length = pl.get("length_km")
    if length is not None:
        rd = _ru_decimal(length)
        ws.cell(22, 12).value = rd
        ws.cell(24, 12).value = rd
        ws.cell(25, 12).value = _ru_decimal(0)

    poles: List[Dict[str, Any]] = list(snapshot_data.get("poles") or [])
    poles.sort(key=lambda p: (p.get("sequence_number") is None, p.get("sequence_number") or 0, str(p.get("pole_number") or "")))

    # Таблица 2.5 — уникальные типы (конструкция / тип опоры) с материалом
    key_pairs = [
        (
            str(p.get("construction") or p.get("pole_type") or "—").strip(),
            str(p.get("material") or "—").strip(),
        )
        for p in poles
    ]
    counts = Counter(key_pairs)
    sorted_items: List[Tuple[Tuple[str, str], int]] = sorted(
        counts.items(),
        key=lambda it: (-it[1], it[0][0], it[0][1]),
    )

    t25_start, t25_end = _find_table_25_bounds(ws)
    _clear_table_25_block(ws, t25_start, t25_end)
    max_rows = max(0, t25_end - t25_start + 1)

    for i, ((ctype, mat), cnt) in enumerate(sorted_items[:max_rows]):
        r = t25_start + i
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 4).value = f"{ctype} ({cnt} шт.)" if cnt > 1 else ctype
        ws.cell(r, 12).value = mat

    total_row = t25_end + 1
    if total_row <= ws.max_row and ws.cell(total_row, 1).value and "ИТОГО" in str(ws.cell(total_row, 1).value).upper():
        ws.cell(total_row, 4).value = f"{len(poles)}(шт.)"

    # Лист с полным поопорным учётом (как ведётся в системе)
    sheet_name = "Поопорный_учёт_авто"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wn = wb.create_sheet(sheet_name)
    hdr = [
        "№ п/п",
        "№ опоры",
        "Порядковый №",
        "Тип опоры",
        "Конструкция",
        "Материал",
        "Высота, м",
        "Год установки",
        "Состояние",
        "Широта",
        "Долгота",
        "Провод (марка)",
        "Сечение",
        "Отпаечная",
        "Примечания",
    ]
    wn.append(hdr)
    for i, p in enumerate(poles, start=1):
        wn.append(
            [
                i,
                p.get("pole_number"),
                p.get("sequence_number"),
                p.get("pole_type"),
                p.get("construction"),
                p.get("material"),
                p.get("height"),
                p.get("year_installed"),
                p.get("condition"),
                p.get("latitude"),
                p.get("longitude"),
                p.get("conductor_type"),
                p.get("conductor_section"),
                "да" if p.get("is_tap_pole") else "нет",
                p.get("notes"),
            ]
        )

    # Сегменты / провода — кратко
    seg_name = "Сегменты_ЛЭП"
    if seg_name in wb.sheetnames:
        del wb[seg_name]
    ws2 = wb.create_sheet(seg_name)
    ws2.append(["mRID", "Наименование", "Длина, км", "кВ", "Отпайка", "Марка провода"])
    for seg in snapshot_data.get("acline_segments") or []:
        ws2.append(
            [
                seg.get("mrid"),
                seg.get("name"),
                seg.get("length_km"),
                seg.get("voltage_level_kv"),
                "да" if seg.get("is_tap") else "нет",
                seg.get("conductor_type"),
            ]
        )

    # Служебный лист: ссылка на СТП и ручные дополнения
    meta_name = "Системные_поля"
    if meta_name in wb.sheetnames:
        del wb[meta_name]
    wm = wb.create_sheet(meta_name)
    wm.append(["Поле", "Значение"])
    wm.append(["Заголовок паспорта", title])
    wm.append(["СТП / норматив", stp_reference or ""])
    if manual_sections:
        for k, v in manual_sections.items():
            wm.append([str(k), json_val(v)])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def json_val(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)
    return str(v)
